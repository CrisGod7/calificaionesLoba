# app/reporter.py
from typing import List, Dict, Any
import pandas as pd
from datetime import datetime
import os


def mostrar_reporte_consola(resultados: List[Dict[str, Any]], mostrar_detalle: bool = True):
    """Muestra los resultados en consola."""
    if not resultados:
        print("No hay resultados para mostrar")
        return

    print("\n" + "="*70)
    print("REPORTE DE CALIFICACIONES".center(70))
    print("="*70)
    print(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"Total de alumnos: {len(resultados)}")
    print("="*70)

    for idx, reporte in enumerate(resultados, 1):
        print(f"\n{'-'*70}")
        print(f"Alumno #{idx}: {reporte['nombre']}")
        print(f"Email: {reporte.get('email', 'N/A')}")
        print(f"Grupo: {reporte.get('grupo', 'N/A')}")
        print(f"{'-'*70}")

        cal_global = reporte.get('calificacion_global', 0)
        porc_global = reporte['porcentaje_global']
        barra_global = crear_barra_progreso(porc_global, longitud=40)
        estado_global = obtener_estado_calificacion(porc_global)

        print(f"Calificacion Global: {cal_global}/10 ({porc_global}%) {estado_global}")
        print(f"   {barra_global}")
        print(f"   Aciertos: {reporte['total_aciertos']}")
        print(f"   Errores: {reporte.get('total_errores', 0)}")
        print(f"   Sin responder: {reporte.get('total_sin_responder', 0)}")

        if mostrar_detalle:
            print(f"\nDesglose por Materia:")
            print(f"{'-'*70}")

            for materia, detalle in sorted(reporte['calificaciones'].items()):
                barra = crear_barra_progreso(detalle['porcentaje'], longitud=25)
                estado = obtener_estado_calificacion(detalle['porcentaje'])

                print(f"\n  {materia}:")
                print(f"     Calificacion: {detalle.get('calificacion', 0)}/10 ({detalle['porcentaje']}%) {estado}")
                print(f"     {barra}")
                print(f"     Aciertos: {detalle['aciertos']}/{detalle['total']} | "
                      f"Errores: {detalle.get('errores', 0)} | "
                      f"Sin resp: {detalle.get('sin_responder', 0)}")

    print(f"\n{'='*70}")
    mostrar_estadisticas_grupo(resultados)
    print(f"{'='*70}\n")


def crear_barra_progreso(porcentaje: float, longitud: int = 30) -> str:
    """Crea una barra de progreso visual."""
    bloques_llenos = int(porcentaje / 100 * longitud)
    bloques_vacios = longitud - bloques_llenos

    if porcentaje >= 80:
        simbolo_lleno = '#'
    elif porcentaje >= 60:
        simbolo_lleno = '='
    else:
        simbolo_lleno = '-'

    return f"[{simbolo_lleno * bloques_llenos}{' ' * bloques_vacios}] {porcentaje}%"


def obtener_estado_calificacion(porcentaje: float) -> str:
    """Retorna un indicador segun el porcentaje."""
    if porcentaje >= 90:
        return "Excelente"
    elif porcentaje >= 80:
        return "Muy bien"
    elif porcentaje >= 70:
        return "Bien"
    elif porcentaje >= 60:
        return "Regular"
    else:
        return "Necesita mejorar"


def mostrar_estadisticas_grupo(resultados: List[Dict[str, Any]]):
    """Muestra estadisticas del grupo."""
    if not resultados:
        return

    print(f"\nESTADISTICAS DEL GRUPO")
    print(f"{'-'*70}")

    promedio_global = sum(r['porcentaje_global'] for r in resultados) / len(resultados)
    promedio_calificacion = sum(r.get('calificacion_global', 0) for r in resultados) / len(resultados)

    excelente = sum(1 for r in resultados if r['porcentaje_global'] >= 90)
    muy_bien = sum(1 for r in resultados if 80 <= r['porcentaje_global'] < 90)
    bien = sum(1 for r in resultados if 70 <= r['porcentaje_global'] < 80)
    regular = sum(1 for r in resultados if 60 <= r['porcentaje_global'] < 70)
    necesita_mejorar = sum(1 for r in resultados if r['porcentaje_global'] < 60)

    print(f"\n  Promedio del grupo: {promedio_calificacion:.2f}/10 ({promedio_global:.2f}%)")
    print(f"\n  Distribucion de calificaciones:")
    print(f"    Excelente (90-100%):        {excelente} alumno(s)")
    print(f"    Muy bien (80-89%):          {muy_bien} alumno(s)")
    print(f"    Bien (70-79%):              {bien} alumno(s)")
    print(f"    Regular (60-69%):           {regular} alumno(s)")
    print(f"    Necesita mejorar (<60%):    {necesita_mejorar} alumno(s)")

    mejor = max(resultados, key=lambda x: x['porcentaje_global'])
    peor = min(resultados, key=lambda x: x['porcentaje_global'])

    print(f"\n  Mejor desempeno: {mejor['nombre']} ({mejor['porcentaje_global']}%)")
    print(f"  Menor desempeno: {peor['nombre']} ({peor['porcentaje_global']}%)")

    print(f"\n  Promedios por materia:")
    if resultados:
        for materia in sorted(resultados[0]['calificaciones'].keys()):
            porcentajes = [r['calificaciones'][materia]['porcentaje'] for r in resultados]
            promedio_materia = sum(porcentajes) / len(porcentajes)
            barra_mini = crear_barra_progreso(promedio_materia, longitud=20)
            print(f"    {materia:15s}: {promedio_materia:5.1f}%  {barra_mini}")


def exportar_a_csv(resultados: List[Dict[str, Any]], ruta_salida: str):
    """Exporta los resultados a CSV con solo aciertos por materia."""
    if not resultados:
        print("No hay resultados para exportar")
        return

    datos_exportar = []

    for reporte in resultados:
        fila = {
            'Nombre': reporte['nombre'],
            'Email': reporte.get('email', ''),
            'Grupo': reporte.get('grupo', ''),
            'Total_Aciertos': reporte['total_aciertos'],
            'Total_Preguntas': reporte['total_preguntas']
        }

        # Solo agregar aciertos por materia (sin porcentajes ni calificaciones)
        for materia, detalle in sorted(reporte['calificaciones'].items()):
            fila[f'{materia}_Aciertos'] = detalle['aciertos']
            fila[f'{materia}_Total'] = detalle['total']

        datos_exportar.append(fila)

    try:
        df = pd.DataFrame(datos_exportar)
        df.to_csv(ruta_salida, index=False, encoding='utf-8-sig')
        print(f"\nResultados exportados a CSV: {ruta_salida}")
        print(f"  Registros: {len(df)}")
        print(f"  Columnas: {len(df.columns)}")
    except Exception as e:
        print(f"Error al exportar CSV: {e}")


def exportar_a_excel(resultados: List[Dict[str, Any]], ruta_salida: str):
    """Exporta los resultados a Excel con solo aciertos."""
    if not resultados:
        print("No hay resultados para exportar")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Calificaciones"

        # Headers simplificados - solo aciertos
        headers = ['Nombre', 'Email', 'Grupo', 'Total Aciertos', 'Total Preguntas']

        if resultados:
            for materia in sorted(resultados[0]['calificaciones'].keys()):
                headers.extend([f'{materia} Aciertos', f'{materia} Total'])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_idx, reporte in enumerate(resultados, 2):
            col = 1
            ws.cell(row=row_idx, column=col, value=reporte['nombre'])
            col += 1
            ws.cell(row=row_idx, column=col, value=reporte.get('email', ''))
            col += 1
            ws.cell(row=row_idx, column=col, value=reporte.get('grupo', ''))
            col += 1
            ws.cell(row=row_idx, column=col, value=reporte['total_aciertos'])
            col += 1
            ws.cell(row=row_idx, column=col, value=reporte['total_preguntas'])
            col += 1

            for materia in sorted(reporte['calificaciones'].keys()):
                detalle = reporte['calificaciones'][materia]
                ws.cell(row=row_idx, column=col, value=detalle['aciertos'])
                col += 1
                ws.cell(row=row_idx, column=col, value=detalle['total'])
                col += 1

        wb.save(ruta_salida)
        print(f"\nResultados exportados a Excel: {ruta_salida}")
    except ImportError:
        print("Para exportar a Excel, instala: pip install openpyxl")
        ruta_csv = ruta_salida.replace('.xlsx', '.csv')
        exportar_a_csv(resultados, ruta_csv)
    except Exception as e:
        print(f"Error al exportar a Excel: {e}")


if __name__ == "__main__":
    print("Modulo reporter.py")
    print("Funciones para generar reportes")