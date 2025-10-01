# app/analisis_errores.py
import pandas as pd
from typing import List, Dict, Any
from datetime import datetime


def generar_reporte_errores_csv(resultados: List[Dict[str, Any]], ruta_salida: str):
    """
    Genera un CSV con las preguntas incorrectas de cada alumno.
    Formato: Una fila por alumno, columnas para cada pregunta (1-110).
    Valor: 0 = correcta, 1 = incorrecta, vacío = sin responder
    """
    if not resultados:
        print("No hay resultados para generar reporte de errores")
        return

    # Obtener el total de preguntas
    total_preguntas = resultados[0]['total_preguntas'] if resultados else 110

    datos_errores = []

    for reporte in resultados:
        fila = {
            'Nombre': reporte['nombre'],
            'Email': reporte.get('email', ''),
            'Grupo': reporte.get('grupo', ''),
            'Total_Aciertos': reporte['total_aciertos'],
            'Total_Errores': reporte.get('total_errores', 0),
            'Total_Sin_Responder': reporte.get('total_sin_responder', 0)
        }

        # Obtener listas de preguntas
        aciertos = set(reporte['estadisticas']['aciertos'])
        errores = set(reporte['estadisticas']['errores'])
        sin_responder = set(reporte['estadisticas']['sin_responder'])

        # Para cada pregunta, marcar su estado
        for num_pregunta in range(1, total_preguntas + 1):
            if num_pregunta in aciertos:
                fila[f'P{num_pregunta}'] = 0  # Correcta
            elif num_pregunta in errores:
                fila[f'P{num_pregunta}'] = 1  # Incorrecta
            elif num_pregunta in sin_responder:
                fila[f'P{num_pregunta}'] = ''  # Sin responder
            else:
                fila[f'P{num_pregunta}'] = ''  # Por defecto

        datos_errores.append(fila)

    try:
        df = pd.DataFrame(datos_errores)
        df.to_csv(ruta_salida, index=False, encoding='utf-8-sig')
        print(f"\nReporte de errores exportado: {ruta_salida}")
        print(f"  Formato: 0=Correcta, 1=Incorrecta, vacio=Sin responder")
    except Exception as e:
        print(f"Error al generar reporte de errores: {e}")


def generar_reporte_errores_por_materia(resultados: List[Dict[str, Any]], ruta_salida: str):
    """
    Genera un CSV con errores agrupados por materia.
    Muestra qué preguntas específicas falló cada alumno en cada materia.
    """
    if not resultados:
        print("No hay resultados")
        return

    datos_materias = []

    for reporte in resultados:
        nombre = reporte['nombre']
        email = reporte.get('email', '')
        grupo = reporte.get('grupo', '')

        # Por cada materia, crear una fila
        for materia, detalle in sorted(reporte['calificaciones'].items()):
            if detalle['errores'] > 0:  # Solo si hay errores
                preguntas_incorrectas = detalle['preguntas_incorrectas']
                preguntas_sin_resp = detalle.get('preguntas_sin_responder', [])

                fila = {
                    'Nombre': nombre,
                    'Email': email,
                    'Grupo': grupo,
                    'Materia': materia,
                    'Aciertos': detalle['aciertos'],
                    'Errores': detalle['errores'],
                    'Sin_Responder': detalle.get('sin_responder', 0),
                    'Total_Preguntas': detalle['total'],
                    'Preguntas_Incorrectas': ', '.join(map(str, preguntas_incorrectas)),
                    'Preguntas_Sin_Responder': ', '.join(map(str, preguntas_sin_resp))
                }
                datos_materias.append(fila)

    try:
        df = pd.DataFrame(datos_materias)
        df.to_csv(ruta_salida, index=False, encoding='utf-8-sig')
        print(f"\nReporte de errores por materia: {ruta_salida}")
    except Exception as e:
        print(f"Error: {e}")


def generar_analisis_preguntas_dificiles(resultados: List[Dict[str, Any]], ruta_salida: str):
    """
    Analiza qué preguntas fueron las más difíciles (más errores).
    Genera un ranking de preguntas con mayor número de errores.
    """
    if not resultados:
        print("No hay resultados")
        return

    # Contar errores por pregunta
    errores_por_pregunta = {}
    total_alumnos = len(resultados)

    for reporte in resultados:
        for pregunta_err in reporte['estadisticas']['errores']:
            errores_por_pregunta[pregunta_err] = errores_por_pregunta.get(pregunta_err, 0) + 1

    # Crear DataFrame
    datos_analisis = []
    for pregunta, num_errores in sorted(errores_por_pregunta.items()):
        porcentaje_error = (num_errores / total_alumnos * 100)
        datos_analisis.append({
            'Pregunta': pregunta,
            'Num_Errores': num_errores,
            'Total_Alumnos': total_alumnos,
            'Porcentaje_Error': round(porcentaje_error, 2)
        })

    # Ordenar por número de errores (descendente)
    df = pd.DataFrame(datos_analisis)
    df = df.sort_values('Num_Errores', ascending=False)

    try:
        df.to_csv(ruta_salida, index=False, encoding='utf-8-sig')
        print(f"\nAnalisis de preguntas dificiles: {ruta_salida}")
        print(f"\nTop 10 preguntas mas dificiles:")
        print(df.head(10).to_string(index=False))
    except Exception as e:
        print(f"Error: {e}")


def generar_matriz_errores_excel(resultados: List[Dict[str, Any]], ruta_salida: str):
    """
    Genera un Excel con una matriz visual de errores.
    Filas = Alumnos, Columnas = Preguntas
    Colores: Verde=Correcto, Rojo=Error, Gris=Sin responder
    """
    if not resultados:
        print("No hay resultados")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz de Errores"

        # Colores
        color_correcto = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        color_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        color_sin_resp = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        color_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        total_preguntas = resultados[0]['total_preguntas'] if resultados else 110

        # Headers
        ws.cell(1, 1, "Nombre").fill = color_header
        ws.cell(1, 1).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, 2, "Email").fill = color_header
        ws.cell(1, 2).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, 3, "Aciertos").fill = color_header
        ws.cell(1, 3).font = Font(bold=True, color="FFFFFF")

        for p in range(1, total_preguntas + 1):
            col = p + 3
            cell = ws.cell(1, col, f"P{p}")
            cell.fill = color_header
            cell.font = Font(bold=True, color="FFFFFF", size=8)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 3

        # Datos
        for row_idx, reporte in enumerate(resultados, 2):
            ws.cell(row_idx, 1, reporte['nombre'])
            ws.cell(row_idx, 2, reporte.get('email', ''))
            ws.cell(row_idx, 3, reporte['total_aciertos'])

            aciertos = set(reporte['estadisticas']['aciertos'])
            errores = set(reporte['estadisticas']['errores'])
            sin_responder = set(reporte['estadisticas']['sin_responder'])

            for p in range(1, total_preguntas + 1):
                col = p + 3
                cell = ws.cell(row_idx, col)

                if p in aciertos:
                    cell.value = "✓"
                    cell.fill = color_correcto
                elif p in errores:
                    cell.value = "✗"
                    cell.fill = color_error
                elif p in sin_responder:
                    cell.value = "-"
                    cell.fill = color_sin_resp

                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=8)

        # Ajustar columnas de info
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10

        wb.save(ruta_salida)
        print(f"\nMatriz visual de errores exportada: {ruta_salida}")
        print("  Verde = Correcta, Rojo = Error, Gris = Sin responder")
    except ImportError:
        print("Para generar matriz visual, instala: pip install openpyxl")
    except Exception as e:
        print(f"Error: {e}")


def generar_todos_reportes_errores(resultados: List[Dict[str, Any]], carpeta_salida: str = 'data/reportes'):
    """
    Genera todos los reportes de errores disponibles.
    """
    import os

    if not os.path.exists(carpeta_salida):
        os.makedirs(carpeta_salida)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    print("\nGenerando reportes de analisis de errores...")
    print("=" * 60)

    # 1. Matriz de errores (0/1 por pregunta)
    ruta1 = os.path.join(carpeta_salida, f'errores_matriz_{timestamp}.csv')
    generar_reporte_errores_csv(resultados, ruta1)

    # 2. Errores por materia
    ruta2 = os.path.join(carpeta_salida, f'errores_por_materia_{timestamp}.csv')
    generar_reporte_errores_por_materia(resultados, ruta2)

    # 3. Análisis de preguntas difíciles
    ruta3 = os.path.join(carpeta_salida, f'preguntas_dificiles_{timestamp}.csv')
    generar_analisis_preguntas_dificiles(resultados, ruta3)

    # 4. Matriz visual en Excel
    ruta4 = os.path.join(carpeta_salida, f'matriz_visual_{timestamp}.xlsx')
    generar_matriz_errores_excel(resultados, ruta4)

    print("=" * 60)
    print(f"Todos los reportes generados en: {carpeta_salida}")


if __name__ == "__main__":
    print("Modulo de analisis de errores")