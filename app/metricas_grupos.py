# app/metricas_grupos.py
import pandas as pd
from typing import List, Dict, Any
from datetime import datetime
from collections import defaultdict


def agrupar_por_grupo(resultados: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """Agrupa los resultados por grupo."""
    grupos = defaultdict(list)

    for resultado in resultados:
        grupo = resultado.get('grupo', 'Sin Grupo')
        if not grupo or str(grupo).lower() in ['nan', 'none', '']:
            grupo = 'Sin Grupo'
        grupos[grupo].append(resultado)

    return dict(grupos)


def calcular_metricas_grupo(alumnos_grupo: List[Dict[str, Any]], nombre_grupo: str) -> Dict[str, Any]:
    """Calcula métricas estadísticas para un grupo específico."""
    if not alumnos_grupo:
        return {}

    total_alumnos = len(alumnos_grupo)

    # Métricas globales del grupo
    aciertos_totales = [a['total_aciertos'] for a in alumnos_grupo]
    promedio_grupo = sum(aciertos_totales) / total_alumnos
    max_aciertos = max(aciertos_totales)
    min_aciertos = min(aciertos_totales)

    mejor_alumno = max(alumnos_grupo, key=lambda x: x['total_aciertos'])
    peor_alumno = min(alumnos_grupo, key=lambda x: x['total_aciertos'])

    # Métricas por materia
    materias = alumnos_grupo[0]['calificaciones'].keys()
    metricas_materias = {}

    for materia in materias:
        aciertos_materia = [a['calificaciones'][materia]['aciertos'] for a in alumnos_grupo]
        total_preguntas = alumnos_grupo[0]['calificaciones'][materia]['total']

        # Prevenir división por cero
        if total_preguntas > 0:
            promedio_materia = sum(aciertos_materia) / total_alumnos
            porcentaje_promedio = (promedio_materia / total_preguntas * 100)
            alumnos_con_100 = sum(1 for a in aciertos_materia if a == total_preguntas)
            alumnos_aprobados = sum(1 for a in aciertos_materia if (a / total_preguntas) * 100 >= 60)
        else:
            # Si la materia tiene 0 preguntas, usar valores por defecto
            promedio_materia = 0
            porcentaje_promedio = 0
            alumnos_con_100 = 0
            alumnos_aprobados = 0

        metricas_materias[materia] = {
            'promedio_aciertos': round(promedio_materia, 2),
            'total_preguntas': total_preguntas,
            'porcentaje_promedio': round(porcentaje_promedio, 2),
            'max_aciertos': max(aciertos_materia) if aciertos_materia else 0,
            'min_aciertos': min(aciertos_materia) if aciertos_materia else 0,
            'alumnos_con_100': alumnos_con_100,
            'alumnos_aprobados': alumnos_aprobados
        }

    # Distribución de calificaciones
    excelente = sum(1 for a in alumnos_grupo if a['porcentaje_global'] >= 90)
    muy_bien = sum(1 for a in alumnos_grupo if 80 <= a['porcentaje_global'] < 90)
    bien = sum(1 for a in alumnos_grupo if 70 <= a['porcentaje_global'] < 80)
    regular = sum(1 for a in alumnos_grupo if 60 <= a['porcentaje_global'] < 70)
    necesita_mejorar = sum(1 for a in alumnos_grupo if a['porcentaje_global'] < 60)

    # Calcular porcentaje promedio del grupo con validación
    total_preguntas_examen = alumnos_grupo[0]['total_preguntas']
    if total_preguntas_examen > 0:
        porcentaje_promedio_grupo = (promedio_grupo / total_preguntas_examen * 100)
    else:
        porcentaje_promedio_grupo = 0

    return {
        'nombre_grupo': nombre_grupo,
        'total_alumnos': total_alumnos,
        'promedio_aciertos': round(promedio_grupo, 2),
        'total_preguntas': total_preguntas_examen,
        'porcentaje_promedio_grupo': round(porcentaje_promedio_grupo, 2),
        'max_aciertos': max_aciertos,
        'min_aciertos': min_aciertos,
        'mejor_alumno': mejor_alumno['nombre'],
        'mejor_alumno_aciertos': mejor_alumno['total_aciertos'],
        'peor_alumno': peor_alumno['nombre'],
        'peor_alumno_aciertos': peor_alumno['total_aciertos'],
        'materias': metricas_materias,
        'distribucion': {
            'excelente_90_100': excelente,
            'muy_bien_80_89': muy_bien,
            'bien_70_79': bien,
            'regular_60_69': regular,
            'necesita_mejorar_0_59': necesita_mejorar
        }
    }


def generar_excel_metricas_grupos(resultados: List[Dict[str, Any]], ruta_salida: str):
    """
    Genera un Excel completo con métricas por grupo.
    Una hoja para resumen general + una hoja por cada grupo.
    """
    if not resultados:
        print("No hay resultados para generar métricas")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Agrupar resultados
        grupos = agrupar_por_grupo(resultados)
        print(f"\nGrupos encontrados: {list(grupos.keys())}")

        # Calcular métricas de cada grupo
        metricas_grupos = {}
        for nombre_grupo, alumnos in grupos.items():
            metricas_grupos[nombre_grupo] = calcular_metricas_grupo(alumnos, nombre_grupo)

        wb = Workbook()

        # Colores
        color_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        color_mejor = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        color_peor = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF")
        font_bold = Font(bold=True)

        # ===== HOJA 1: RESUMEN GENERAL =====
        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"

        # Título
        ws_resumen['A1'] = 'COMPARACION ENTRE GRUPOS'
        ws_resumen['A1'].font = Font(bold=True, size=14)
        ws_resumen.merge_cells('A1:F1')

        # Headers
        headers = ['Grupo', 'Total Alumnos', 'Promedio Aciertos', 'Total Preguntas',
                   'Porcentaje', 'Mejor Alumno']
        for col, header in enumerate(headers, 1):
            cell = ws_resumen.cell(3, col, header)
            cell.fill = color_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center")

        # Datos de cada grupo
        grupos_ordenados = sorted(metricas_grupos.items(),
                                  key=lambda x: x[1]['porcentaje_promedio_grupo'],
                                  reverse=True)

        for row, (nombre_grupo, metricas) in enumerate(grupos_ordenados, 4):
            ws_resumen.cell(row, 1, nombre_grupo)
            ws_resumen.cell(row, 2, metricas['total_alumnos'])
            ws_resumen.cell(row, 3, metricas['promedio_aciertos'])
            ws_resumen.cell(row, 4, metricas['total_preguntas'])
            ws_resumen.cell(row, 5, f"{metricas['porcentaje_promedio_grupo']}%")
            ws_resumen.cell(row, 6, f"{metricas['mejor_alumno']} ({metricas['mejor_alumno_aciertos']})")

            # Resaltar mejor y peor grupo
            if row == 4:  # Mejor grupo
                for col in range(1, 7):
                    ws_resumen.cell(row, col).fill = color_mejor
            elif row == 3 + len(grupos_ordenados):  # Peor grupo
                for col in range(1, 7):
                    ws_resumen.cell(row, col).fill = color_peor

        # Comparación por materia
        row_materia = len(grupos_ordenados) + 6
        ws_resumen.cell(row_materia, 1, 'COMPARACION POR MATERIA').font = font_bold
        ws_resumen.merge_cells(f'A{row_materia}:E{row_materia}')

        row_materia += 2
        materias = list(metricas_grupos[list(grupos.keys())[0]]['materias'].keys())

        # Headers de materias
        ws_resumen.cell(row_materia, 1, 'Materia').fill = color_header
        ws_resumen.cell(row_materia, 1).font = font_header
        for col, grupo in enumerate(sorted(grupos.keys()), 2):
            ws_resumen.cell(row_materia, col, grupo).fill = color_header
            ws_resumen.cell(row_materia, col).font = font_header

        # Datos por materia
        for materia in sorted(materias):
            row_materia += 1
            ws_resumen.cell(row_materia, 1, materia)
            for col, grupo in enumerate(sorted(grupos.keys()), 2):
                porcentaje = metricas_grupos[grupo]['materias'][materia]['porcentaje_promedio']
                ws_resumen.cell(row_materia, col, f"{porcentaje}%")

        # Ajustar anchos
        for col in range(1, 7):
            ws_resumen.column_dimensions[get_column_letter(col)].width = 18

        # ===== HOJAS POR GRUPO =====
        for nombre_grupo, alumnos in sorted(grupos.items()):
            ws_grupo = wb.create_sheet(f"Grupo {nombre_grupo}")
            metricas = metricas_grupos[nombre_grupo]

            # Título del grupo
            ws_grupo['A1'] = f'METRICAS DEL GRUPO: {nombre_grupo}'
            ws_grupo['A1'].font = Font(bold=True, size=14)
            ws_grupo.merge_cells('A1:E1')

            # Información general
            ws_grupo['A3'] = 'Total de alumnos:'
            ws_grupo['B3'] = metricas['total_alumnos']
            ws_grupo['A4'] = 'Promedio de aciertos:'
            ws_grupo['B4'] = metricas['promedio_aciertos']
            ws_grupo['A5'] = 'Porcentaje promedio:'
            ws_grupo['B5'] = f"{metricas['porcentaje_promedio_grupo']}%"
            ws_grupo['A6'] = 'Mejor alumno:'
            ws_grupo['B6'] = f"{metricas['mejor_alumno']} ({metricas['mejor_alumno_aciertos']} aciertos)"
            ws_grupo['A7'] = 'Alumno con mas area de oportunidad:'
            ws_grupo['B7'] = f"{metricas['peor_alumno']} ({metricas['peor_alumno_aciertos']} aciertos)"

            # Distribución
            ws_grupo['A9'] = 'DISTRIBUCION DE CALIFICACIONES'
            ws_grupo['A9'].font = font_bold
            dist = metricas['distribucion']
            ws_grupo['A10'] = 'Excelente (90-100%):'
            ws_grupo['B10'] = dist['excelente_90_100']
            ws_grupo['A11'] = 'Muy bien (80-89%):'
            ws_grupo['B11'] = dist['muy_bien_80_89']
            ws_grupo['A12'] = 'Bien (70-79%):'
            ws_grupo['B12'] = dist['bien_70_79']
            ws_grupo['A13'] = 'Regular (60-69%):'
            ws_grupo['B13'] = dist['regular_60_69']
            ws_grupo['A14'] = 'Necesita mejorar (<60%):'
            ws_grupo['B14'] = dist['necesita_mejorar_0_59']

            # Métricas por materia
            ws_grupo['A16'] = 'METRICAS POR MATERIA'
            ws_grupo['A16'].font = font_bold
            ws_grupo.merge_cells('A16:F16')

            headers_materia = ['Materia', 'Promedio', 'Total Preguntas', 'Porcentaje',
                               'Max', 'Alumnos 100%']
            for col, header in enumerate(headers_materia, 1):
                cell = ws_grupo.cell(17, col, header)
                cell.fill = color_header
                cell.font = font_header

            row = 18
            for materia, datos in sorted(metricas['materias'].items()):
                ws_grupo.cell(row, 1, materia)
                ws_grupo.cell(row, 2, datos['promedio_aciertos'])
                ws_grupo.cell(row, 3, datos['total_preguntas'])
                ws_grupo.cell(row, 4, f"{datos['porcentaje_promedio']}%")
                ws_grupo.cell(row, 5, datos['max_aciertos'])
                ws_grupo.cell(row, 6, datos['alumnos_con_100'])
                row += 1

            # Lista de alumnos
            row += 2
            ws_grupo.cell(row, 1, 'LISTA DE ALUMNOS').font = font_bold
            ws_grupo.merge_cells(f'A{row}:D{row}')

            row += 1
            headers_alumnos = ['Nombre', 'Email', 'Total Aciertos', 'Porcentaje']
            for col, header in enumerate(headers_alumnos, 1):
                cell = ws_grupo.cell(row, col, header)
                cell.fill = color_header
                cell.font = font_header

            row += 1
            for alumno in sorted(alumnos, key=lambda x: x['total_aciertos'], reverse=True):
                ws_grupo.cell(row, 1, alumno['nombre'])
                ws_grupo.cell(row, 2, alumno.get('email', ''))
                ws_grupo.cell(row, 3, alumno['total_aciertos'])
                ws_grupo.cell(row, 4, f"{alumno['porcentaje_global']}%")
                row += 1

            # Ajustar anchos
            for col in range(1, 7):
                ws_grupo.column_dimensions[get_column_letter(col)].width = 20

        wb.save(ruta_salida)
        print(f"\nMetricas por grupo exportadas: {ruta_salida}")
        print(f"  Hojas generadas: Resumen General + {len(grupos)} grupos")

    except ImportError:
        print("Para generar Excel, instala: pip install openpyxl")
    except Exception as e:
        print(f"Error al generar metricas: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    print("Modulo de metricas por grupo")