# app/excel_consolidado.py
"""
Generador de reportes consolidados en un único archivo Excel.
Cada tipo de reporte se almacena en una hoja diferente.
"""

import pandas as pd
from typing import List, Dict, Any
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


class GeneradorExcelConsolidado:
    """Genera un archivo Excel único con múltiples hojas de reportes."""

    def __init__(self, resultados: List[Dict[str, Any]]):
        self.resultados = resultados
        self.wb = Workbook()

        # Colores del tema Lobatchewsky
        self.COLOR_PRINCIPAL = "005DAB"  # Azul corporativo
        self.COLOR_HEADER = "003D7A"  # Azul más oscuro
        self.COLOR_EXCELENTE = "C6EFCE"  # Verde claro
        self.COLOR_BIEN = "FFEB9C"  # Amarillo claro
        self.COLOR_REGULAR = "FFC7CE"  # Rojo claro
        self.COLOR_BLANCO = "FFFFFF"
        self.COLOR_GRIS = "F2F2F2"

    def generar_reporte_completo(self, ruta_salida: str):
        """Genera el archivo Excel consolidado con todas las hojas."""
        print("\n" + "=" * 70)
        print("📊 GENERANDO REPORTE EXCEL CONSOLIDADO")
        print("=" * 70)

        # Eliminar hoja por defecto
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # 1. Hoja de Resumen General
        print("  ✓ Generando Resumen General...")
        self._crear_hoja_resumen()

        # 2. Hoja de Calificaciones Detalladas
        print("  ✓ Generando Calificaciones Detalladas...")
        self._crear_hoja_calificaciones()

        # 3. Hojas por Grupo
        grupos = self._agrupar_por_grupo()
        for nombre_grupo in sorted(grupos.keys()):
            print(f"  ✓ Generando hoja para Grupo {nombre_grupo}...")
            self._crear_hoja_grupo(nombre_grupo, grupos[nombre_grupo])

        # 4. Hoja de Análisis de Errores
        print("  ✓ Generando Análisis de Errores...")
        self._crear_hoja_errores()

        # 5. Hoja de Preguntas Difíciles
        print("  ✓ Generando Análisis de Preguntas Difíciles...")
        self._crear_hoja_preguntas_dificiles()

        # 6. Hoja de Matriz Visual
        print("  ✓ Generando Matriz Visual de Errores...")
        self._crear_hoja_matriz_visual()

        # Guardar archivo
        self.wb.save(ruta_salida)
        print("=" * 70)
        print(f"✅ Reporte consolidado generado: {ruta_salida}")
        print(f"   Total de hojas: {len(self.wb.sheetnames)}")
        print("=" * 70 + "\n")

    def _crear_hoja_resumen(self):
        """Crea la hoja de resumen general."""
        ws = self.wb.create_sheet("📋 Resumen General", 0)

        # Título principal
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = 'CLUB DE MATEMÁTICAS LOBATCHEWSKY'
        cell.font = Font(size=16, bold=True, color=self.COLOR_BLANCO)
        cell.fill = PatternFill(start_color=self.COLOR_PRINCIPAL,
                                end_color=self.COLOR_PRINCIPAL,
                                fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Subtítulo
        ws.merge_cells('A2:F2')
        cell = ws['A2']
        cell.value = f'Reporte General de Calificaciones - {datetime.now().strftime("%d/%m/%Y")}'
        cell.font = Font(size=12, italic=True)
        cell.alignment = Alignment(horizontal="center")

        # Estadísticas principales
        row = 4
        ws[f'A{row}'] = 'ESTADÍSTICAS PRINCIPALES'
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.COLOR_PRINCIPAL)
        ws.merge_cells(f'A{row}:F{row}')

        row += 2
        estadisticas = self._calcular_estadisticas_generales()

        datos_resumen = [
            ('Total de Alumnos', estadisticas['total_alumnos']),
            ('Promedio General', f"{estadisticas['promedio_global']:.2f}%"),
            ('Calificación Promedio', f"{estadisticas['calificacion_promedio']:.2f}/10"),
            ('Mejor Desempeño', f"{estadisticas['mejor_alumno']} ({estadisticas['mejor_porcentaje']:.1f}%)"),
            ('Menor Desempeño', f"{estadisticas['peor_alumno']} ({estadisticas['peor_porcentaje']:.1f}%)"),
        ]

        for label, valor in datos_resumen:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'D{row}'] = valor
            ws[f'D{row}'].alignment = Alignment(horizontal="left")
            row += 1

        # Distribución de calificaciones
        row += 2
        ws[f'A{row}'] = 'DISTRIBUCIÓN DE CALIFICACIONES'
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.COLOR_PRINCIPAL)
        ws.merge_cells(f'A{row}:F{row}')

        row += 2
        headers = ['Rango', 'Cantidad', 'Porcentaje', 'Barra']
        for col, header in enumerate(headers, 2):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color=self.COLOR_BLANCO)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row += 1
        distribucion = estadisticas['distribucion']
        for rango, cantidad in distribucion.items():
            porcentaje = (cantidad / estadisticas['total_alumnos'] * 100)
            ws.cell(row, 2, rango)
            ws.cell(row, 3, cantidad)
            ws.cell(row, 4, f"{porcentaje:.1f}%")

            # Barra visual
            barra = '█' * int(porcentaje / 5)
            ws.cell(row, 5, barra)
            row += 1

        # Promedio por materia
        row += 2
        ws[f'A{row}'] = 'PROMEDIO POR MATERIA'
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.COLOR_PRINCIPAL)
        ws.merge_cells(f'A{row}:F{row}')

        row += 2
        headers = ['Materia', 'Promedio', 'Aciertos Prom.', 'Total Preguntas']
        for col, header in enumerate(headers, 2):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color=self.COLOR_BLANCO)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type="solid")

        row += 1
        for materia, datos in sorted(estadisticas['materias'].items()):
            ws.cell(row, 2, materia)
            ws.cell(row, 3, f"{datos['promedio']:.1f}%")
            ws.cell(row, 4, f"{datos['aciertos_prom']:.1f}")
            ws.cell(row, 5, datos['total'])

            # Color según rendimiento
            color = self._obtener_color_rendimiento(datos['promedio'])
            for col in range(2, 6):
                ws.cell(row, col).fill = PatternFill(start_color=color,
                                                     end_color=color,
                                                     fill_type="solid")
            row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 5

    def _crear_hoja_calificaciones(self):
        """Crea la hoja con todas las calificaciones detalladas."""
        ws = self.wb.create_sheet("📝 Calificaciones")

        # Headers
        headers = ['#', 'Nombre', 'Email', 'Grupo', 'Total', 'Calificación', '%']

        # Agregar materias
        if self.resultados:
            materias = sorted(self.resultados[0]['calificaciones'].keys())
            for materia in materias:
                headers.append(f'{materia}\nAciertos')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color=self.COLOR_BLANCO, size=11)
            cell.fill = PatternFill(start_color=self.COLOR_PRINCIPAL,
                                    end_color=self.COLOR_PRINCIPAL,
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center",
                                       wrap_text=True)

        ws.row_dimensions[1].height = 35

        # Datos
        for idx, resultado in enumerate(self.resultados, 2):
            ws.cell(idx, 1, idx - 1)
            ws.cell(idx, 2, resultado['nombre'])
            ws.cell(idx, 3, resultado.get('email', ''))
            ws.cell(idx, 4, resultado.get('grupo', ''))
            ws.cell(idx, 5, f"{resultado['total_aciertos']}/{resultado['total_preguntas']}")
            ws.cell(idx, 6, f"{resultado['calificacion_global']:.2f}")
            ws.cell(idx, 7, f"{resultado['porcentaje_global']:.1f}%")

            # Colorear calificación según rendimiento
            color = self._obtener_color_rendimiento(resultado['porcentaje_global'])
            for col in range(1, 8):
                ws.cell(idx, col).fill = PatternFill(start_color=color,
                                                     end_color=color,
                                                     fill_type="solid")

            # Materias
            col = 8
            for materia in sorted(resultado['calificaciones'].keys()):
                datos = resultado['calificaciones'][materia]
                ws.cell(idx, col, f"{datos['aciertos']}/{datos['total']}")
                ws.cell(idx, col).alignment = Alignment(horizontal="center")
                col += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10

        for col in range(8, 8 + len(materias)):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _crear_hoja_grupo(self, nombre_grupo: str, alumnos: List[Dict[str, Any]]):
        """Crea una hoja específica para un grupo."""
        # Limitar nombre de hoja (Excel tiene límite de 31 caracteres)
        nombre_hoja = f"Grupo {nombre_grupo}"[:31]
        ws = self.wb.create_sheet(nombre_hoja)

        # Título
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = f'GRUPO: {nombre_grupo}'
        cell.font = Font(size=14, bold=True, color=self.COLOR_BLANCO)
        cell.fill = PatternFill(start_color=self.COLOR_PRINCIPAL,
                                end_color=self.COLOR_PRINCIPAL,
                                fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        # Estadísticas del grupo
        row = 3
        promedio_grupo = sum(a['porcentaje_global'] for a in alumnos) / len(alumnos)

        ws[f'A{row}'] = 'Total de Alumnos:'
        ws[f'B{row}'] = len(alumnos)
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Promedio del Grupo:'
        ws[f'B{row}'] = f"{promedio_grupo:.2f}%"
        ws[f'A{row}'].font = Font(bold=True)

        # Lista de alumnos
        row += 3
        headers = ['#', 'Nombre', 'Total', 'Calificación', '%']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color=self.COLOR_BLANCO)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row += 1
        for idx, alumno in enumerate(sorted(alumnos,
                                            key=lambda x: x['porcentaje_global'],
                                            reverse=True), 1):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, alumno['nombre'])
            ws.cell(row, 3, f"{alumno['total_aciertos']}/{alumno['total_preguntas']}")
            ws.cell(row, 4, f"{alumno['calificacion_global']:.2f}")
            ws.cell(row, 5, f"{alumno['porcentaje_global']:.1f}%")

            color = self._obtener_color_rendimiento(alumno['porcentaje_global'])
            for col in range(1, 6):
                ws.cell(row, col).fill = PatternFill(start_color=color,
                                                     end_color=color,
                                                     fill_type="solid")
            row += 1

        # Promedio por materia del grupo
        row += 2
        ws[f'A{row}'] = 'Promedio por Materia'
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLOR_PRINCIPAL)
        ws.merge_cells(f'A{row}:E{row}')

        row += 2
        headers = ['Materia', 'Promedio', 'Aciertos Prom.', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color=self.COLOR_BLANCO)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type="solid")

        row += 1
        materias = sorted(alumnos[0]['calificaciones'].keys())
        for materia in materias:
            aciertos = [a['calificaciones'][materia]['aciertos'] for a in alumnos]
            total_preg = alumnos[0]['calificaciones'][materia]['total']
            promedio = sum(aciertos) / len(aciertos)
            porcentaje = (promedio / total_preg * 100) if total_preg > 0 else 0

            ws.cell(row, 1, materia)
            ws.cell(row, 2, f"{porcentaje:.1f}%")
            ws.cell(row, 3, f"{promedio:.1f}")
            ws.cell(row, 4, total_preg)

            color = self._obtener_color_rendimiento(porcentaje)
            for col in range(1, 5):
                ws.cell(row, col).fill = PatternFill(start_color=color,
                                                     end_color=color,
                                                     fill_type="solid")
            row += 1

        # Ajustar anchos
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _crear_hoja_errores(self):
        """Crea hoja con análisis de errores por alumno."""
        ws = self.wb.create_sheet("❌ Análisis Errores")

        # Headers
        headers = ['Nombre', 'Email', 'Grupo', 'Total Aciertos',
                   'Total Errores', 'Sin Responder']

        # Agregar columnas para preguntas
        total_preguntas = self.resultados[0]['total_preguntas']
        for p in range(1, total_preguntas + 1):
            headers.append(f'P{p}')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color=self.COLOR_BLANCO, size=9)
            cell.fill = PatternFill(start_color=self.COLOR_PRINCIPAL,
                                    end_color=self.COLOR_PRINCIPAL,
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for idx, resultado in enumerate(self.resultados, 2):
            ws.cell(idx, 1, resultado['nombre'])
            ws.cell(idx, 2, resultado.get('email', ''))
            ws.cell(idx, 3, resultado.get('grupo', ''))
            ws.cell(idx, 4, resultado['total_aciertos'])
            ws.cell(idx, 5, resultado['total_errores'])
            ws.cell(idx, 6, resultado['total_sin_responder'])

            aciertos = set(resultado['estadisticas']['aciertos'])
            errores = set(resultado['estadisticas']['errores'])
            sin_resp = set(resultado['estadisticas']['sin_responder'])

            for p in range(1, total_preguntas + 1):
                col = 6 + p
                if p in aciertos:
                    cell = ws.cell(idx, col, '✓')
                    cell.fill = PatternFill(start_color=self.COLOR_EXCELENTE,
                                            end_color=self.COLOR_EXCELENTE,
                                            fill_type="solid")
                elif p in errores:
                    cell = ws.cell(idx, col, '✗')
                    cell.fill = PatternFill(start_color=self.COLOR_REGULAR,
                                            end_color=self.COLOR_REGULAR,
                                            fill_type="solid")
                elif p in sin_resp:
                    cell = ws.cell(idx, col, '-')
                    cell.fill = PatternFill(start_color=self.COLOR_GRIS,
                                            end_color=self.COLOR_GRIS,
                                            fill_type="solid")

                ws.cell(idx, col).alignment = Alignment(horizontal="center")
                ws.cell(idx, col).font = Font(size=8)

        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        for col in range(4, 7):
            ws.column_dimensions[get_column_letter(col)].width = 12
        for col in range(7, 7 + total_preguntas):
            ws.column_dimensions[get_column_letter(col)].width = 3

    def _crear_hoja_preguntas_dificiles(self):
        """Crea hoja con análisis de preguntas difíciles."""
        ws = self.wb.create_sheet("📊 Preguntas Difíciles")

        # Calcular errores por pregunta
        errores_por_pregunta = defaultdict(int)
        total_alumnos = len(self.resultados)

        for resultado in self.resultados:
            for pregunta in resultado['estadisticas']['errores']:
                errores_por_pregunta[pregunta] += 1

        # Título
        ws['A1'] = 'ANÁLISIS DE PREGUNTAS DIFÍCILES'
        ws['A1'].font = Font(size=14, bold=True, color=self.COLOR_BLANCO)
        ws['A1'].fill = PatternFill(start_color=self.COLOR_PRINCIPAL,
                                    end_color=self.COLOR_PRINCIPAL,
                                    fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:E1')

        # Headers
        headers = ['Pregunta', 'Número Errores', 'Total Alumnos',
                   '% Error', 'Dificultad']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True, color=self.COLOR_BLANCO)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Datos ordenados por dificultad
        row = 4
        for pregunta, num_errores in sorted(errores_por_pregunta.items(),
                                            key=lambda x: x[1],
                                            reverse=True):
            porcentaje_error = (num_errores / total_alumnos * 100)

            if porcentaje_error >= 70:
                dificultad = "Muy Difícil"
                color = self.COLOR_REGULAR
            elif porcentaje_error >= 50:
                dificultad = "Difícil"
                color = self.COLOR_BIEN
            else:
                dificultad = "Normal"
                color = self.COLOR_EXCELENTE

            ws.cell(row, 1, f"P{pregunta}")
            ws.cell(row, 2, num_errores)
            ws.cell(row, 3, total_alumnos)
            ws.cell(row, 4, f"{porcentaje_error:.1f}%")
            ws.cell(row, 5, dificultad)

            for col in range(1, 6):
                ws.cell(row, col).fill = PatternFill(start_color=color,
                                                     end_color=color,
                                                     fill_type="solid")
                ws.cell(row, col).alignment = Alignment(horizontal="center")

            row += 1

        # Ajustar anchos
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _crear_hoja_matriz_visual(self):
        """Crea matriz visual simplificada de errores."""
        ws = self.wb.create_sheet("🔲 Matriz Visual")

        # Similar a la hoja de errores pero más compacta
        # (Reutiliza lógica de _crear_hoja_errores pero más visual)
        self._crear_hoja_errores()  # Simplificado por ahora

    # Métodos auxiliares

    def _agrupar_por_grupo(self) -> Dict[str, List[Dict[str, Any]]]:
        """Agrupa resultados por grupo."""
        grupos = defaultdict(list)
        for resultado in self.resultados:
            grupo = resultado.get('grupo', 'Sin Grupo')
            if not grupo or str(grupo).lower() in ['nan', 'none', '']:
                grupo = 'Sin Grupo'
            grupos[grupo].append(resultado)
        return dict(grupos)

    def _calcular_estadisticas_generales(self) -> Dict[str, Any]:
        """Calcula estadísticas generales de todos los alumnos."""
        total = len(self.resultados)
        promedio = sum(r['porcentaje_global'] for r in self.resultados) / total
        calificacion = sum(r['calificacion_global'] for r in self.resultados) / total

        mejor = max(self.resultados, key=lambda x: x['porcentaje_global'])
        peor = min(self.resultados, key=lambda x: x['porcentaje_global'])

        # Distribución
        distribucion = {
            'Excelente (90-100%)': sum(1 for r in self.resultados if r['porcentaje_global'] >= 90),
            'Muy bien (80-89%)': sum(1 for r in self.resultados if 80 <= r['porcentaje_global'] < 90),
            'Bien (70-79%)': sum(1 for r in self.resultados if 70 <= r['porcentaje_global'] < 80),
            'Regular (60-69%)': sum(1 for r in self.resultados if 60 <= r['porcentaje_global'] < 70),
            'Necesita mejorar (<60%)': sum(1 for r in self.resultados if r['porcentaje_global'] < 60)
        }

        # Promedio por materia
        materias = {}
        for materia in self.resultados[0]['calificaciones'].keys():
            aciertos = [r['calificaciones'][materia]['aciertos'] for r in self.resultados]
            total_preg = self.resultados[0]['calificaciones'][materia]['total']
            promedio_aciertos = sum(aciertos) / len(aciertos)
            porcentaje = (promedio_aciertos / total_preg * 100) if total_preg > 0 else 0

            materias[materia] = {
                'promedio': porcentaje,
                'aciertos_prom': promedio_aciertos,
                'total': total_preg
            }

        return {
            'total_alumnos': total,
            'promedio_global': promedio,
            'calificacion_promedio': calificacion,
            'mejor_alumno': mejor['nombre'],
            'mejor_porcentaje': mejor['porcentaje_global'],
            'peor_alumno': peor['nombre'],
            'peor_porcentaje': peor['porcentaje_global'],
            'distribucion': distribucion,
            'materias': materias
        }

    def _obtener_color_rendimiento(self, porcentaje: float) -> str:
        """Retorna color según el porcentaje de rendimiento."""
        if porcentaje >= 80:
            return self.COLOR_EXCELENTE
        elif porcentaje >= 60:
            return self.COLOR_BIEN
        else:
            return self.COLOR_REGULAR


def generar_reporte_consolidado(resultados: List[Dict[str, Any]],
                                ruta_salida: str):
    """
    Función principal para generar el reporte consolidado.

    Args:
        resultados: Lista de resultados de calificaciones
        ruta_salida: Ruta donde guardar el archivo Excel
    """
    generador = GeneradorExcelConsolidado(resultados)
    generador.generar_reporte_completo(ruta_salida)


if __name__ == "__main__":
    print("Módulo de generación de reportes consolidados")